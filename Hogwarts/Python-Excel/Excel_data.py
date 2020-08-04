
from openpyxl import Workbook

wb = Workbook()
#抓取活动工作表
ws = wb.active
ws['A1'] = '身高'
ws['B1'] = '体重'

#身高
height = [180,175,170,174,169]
#体重
weight = [80,74,65,67,60]

for i in range(len(height)):
    print(i)
    ws.cell(row=2+i,column=1,value=height[i])
    ws.cell(row=2+i, column=2,value=weight[i])

#保存文件
wb.save('sample.xlsx')