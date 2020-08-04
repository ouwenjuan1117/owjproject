
from openpyxl import Workbook,load_workbook



class PracticeExcel(object):
    #定义一个写的方法
    def create(self):
        wb = Workbook()
        # 抓取活动工作表
        ws = wb.active
        ws['A1'] = '身高'
        ws['B1'] = '体重'
        # 身高
        self.height = [180, 175, 170, 174, 169]
        # 体重
        self.weight = [70, 77, 48, 55, 80]

        for i in range(len(self.height)):
            ws.cell(row=2 + i, column=1, value=self.height[i])
            ws.cell(row=2 + i, column=2, value=self.weight[i])

        # 保存文件
        wb.save('sample2.xlsx')

    def health_weight(self):
        ld = load_workbook(filename='sample2.xlsx')
        sheet = ld.active
        sheet['C1'] ='BMI'
        for i in range(len(self.height)):
            height = sheet.cell(row=2+i, column=1).value
            weight = sheet.cell(row=2+i, column=2).value
            # print(height)
            # print(weight)
            #获取身高对应的健康体重
            health_w =weight/((height/100)**2)
            if 18<health_w<25:
                print(health_w)
                sheet.cell(row=2+i, column=3).value = "正常体重"
            elif health_w<18:
                print(health_w)
                sheet.cell(row=2+i, column=3).value = "偏瘦"
            elif 25<health_w<30:
                print(health_w)
                sheet.cell(row=2+i, column=3).value = "微胖"
            elif health_w>30:
                sheet.cell(row=2+i, column=3).value = "轻度肥胖"
            elif health_w>35:
                sheet.cell(row=2+i, column=3).value = "中度肥胖"
            elif health_w > 40:
                sheet.cell(row=2 + i, column=3).value = "重度肥胖"

        # 注意，一定要保存
        ld.save("sample2.xlsx")
if __name__=='__main__':
    pe = PracticeExcel()
    pe.create()
    pe.health_weight()
